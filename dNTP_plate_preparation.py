path = r'C:\Users\Ludivine Profit\Desktop\Synthesis development\Hamilton\Hamilton script\Reagent preparation\dNTP_plate_design.xlsx'
outputPath=r'C:\Users\Ludivine Profit\Desktop\Synthesis development\Hamilton\Hamilton script\Reagent preparation\dNTP_plate_preparation.xlsx'
import openpyxl

def getExcelSheet(path):
    """
    Returns the first sheet of the excel at the path
    :param path:
    :return:
    """
    dNTP_excel = openpyxl.load_workbook(path,data_only=True)
    dNTP_sheet = dNTP_excel["Feuil1"]
    return dNTP_sheet

def getPlateOrder(dNTP_sheet):
    plateOrder =[]
    plateOrderIndexes = []
    for row in range(1, 100):
        for col in range(1, 100):
            value = dNTP_sheet.cell(row, col).value
            if value == "Plate order":
                plateOrderIndexes = [row, col]
    for col in range (1,60):
        plateNbr = dNTP_sheet.cell(plateOrderIndexes[0],plateOrderIndexes[1]+col).value
        if plateNbr != "":
            plateOrder.append(plateNbr)
    return plateOrder

def getSequences(dNTP_sheet):
    """
    Reads the first sheet of the Quartet Control file and returns all the sequences in it in an array
    :param dNTP_sheet: an excel sheet
    :return:
    """
    sequences=[]

    sequenceLayoutIndexes = []
    for row in range(1, 100):
        for col in range(1, 100):
            value = dNTP_sheet.cell(row, col).value
            if value == "Sequence layout":
                sequenceLayoutIndexes = [row, col]

    for col in range(1,13):
        for row in range(1,9):
            sequence=str(dNTP_sheet.cell(sequenceLayoutIndexes[0]+row,sequenceLayoutIndexes[1]+col).value)
            sequences.append(sequence)

    return sequences

def getConcentrations(dNTP_sheet):
    """
    Reads the first sheet of the Quartet Control file and returns all the sequences in it in an array
    :param nucs_sheet: an excel sheet
    :return:
    """

    concentrations=[]

    concentrationsIndexes = []
    for row in range(1, 100):
        for col in range(1, 100):
            value = dNTP_sheet.cell(row, col).value
            if value == "Concentrations":
                concentrationsIndexes = [row, col]

    for col in range(1,13):
        for row in range(1,9):
            concentration=dNTP_sheet.cell(concentrationsIndexes[0]+row,concentrationsIndexes[1]+col).value
            concentrations.append(concentration)

    return concentrations

def getVolumes(dNTP_sheet):
    """
       Reads the first sheet of the Quartet Control file and returns all the sequences in it in an array
       :param nucs_sheet: an excel sheet
       :return:
       """

    volumes = []

    volumesIndexes = []
    for row in range(1, 100):
        for col in range(1, 100):
            value = dNTP_sheet.cell(row, col).value
            if value == "Volume/well (µL)":
                volumesIndexes = [row, col]
    print("volumes indexes =")
    print(volumesIndexes)

    for plate in range(1,5):
            volume = dNTP_sheet.cell(volumesIndexes[0],volumesIndexes[1]+plate).value
            volumes.append(volume)
    return volumes


def getPlates(sequences,plateOrder,concentrations):
    #plates = array of 96 wells for 4 nuc plates for 4 premix plates
    #x = number of well
    x = 96
    plates=[[[0 for well in range (x)]for nucPlates in range (4)]for premixPlates in range (4)]
    well = 0
    for sequence in sequences:
        i = 0
        for nuc in sequence:

                #print("well = ")
                #print(well)
                #print("plateOrder = ")
                #print(plateOrder[i])
                #print("nuc = ")
                #print(nuc)
                if nuc == "A":
                    plates[plateOrder[i]-1][0][well] = concentrations[well]
                if nuc == "C":
                    plates[plateOrder[i]-1][1][well]= concentrations[well]
                if nuc == "G":
                    plates[plateOrder[i]-1][2][well]= concentrations[well]
                if nuc == "T":
                    plates[plateOrder[i]-1][3][well]= concentrations[well]
                i = i + 1
                #print("plate = ")
                #print(plates)
        well = well + 1
    return plates

def writeExcel(plates):
    Plate_preparation_excel = openpyxl.load_workbook(outputPath)
    sheet = Plate_preparation_excel["Plates"]
    offsetRow=12
    offsetCol=3

    for premixPlate in range (4):
        sheet.cell(5,2+premixPlate*15).value = volumes[premixPlate]

    for premixPlate in range (4):
        for nucPlate in range (4):
            well=0
            for col in range (12):
                for row in range (8):
                    sheet.cell(offsetRow+row, offsetCol+col).value = plates[premixPlate][nucPlate][well]
                    well = well + 1
            offsetRow = offsetRow + 14
        offsetRow=12
        offsetCol = offsetCol + 15

    Plate_preparation_excel.save(outputPath)

nucs_sheet=getExcelSheet(path)
plateOrder=getPlateOrder(nucs_sheet)
print("plateOrder = ")
print(plateOrder)
sequences=getSequences(nucs_sheet)
print("sequences = ")
print(sequences)
concentrations=getConcentrations(nucs_sheet)
print("concentrations = ")
print(concentrations)
plates=getPlates(sequences,plateOrder,concentrations)
print("volumes =")
volumes=getVolumes(nucs_sheet)
print(volumes)
print("plates = ")
print(plates)
writeExcel(plates)
