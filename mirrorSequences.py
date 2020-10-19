
path = r'C:\Users\Ludivine Profit\Desktop\Synthesis development\Hamilton\Hamilton script\Reagent preparation\dNTP_plate_design.xlsx'
outputPath = r'C:\Users\Ludivine Profit\Desktop\Synthesis development\Hamilton\Hamilton script\Reagent preparation\Mirror_sequences.xlsx'
import openpyxl

def getMotherSequences(path):
    dNTP_excel = openpyxl.load_workbook(path)
    dNTP_sheet = dNTP_excel["Feuil1"]
    motherSequenceIndexes = []
    for row in range(1,100):
        for col in range(1,100):
            value = dNTP_sheet.cell(row,col).value
            if value == "Mother sequence":
                motherSequenceIndexes = [row,col]
    sequence = dNTP_sheet.cell(motherSequenceIndexes[0], motherSequenceIndexes[1]+1).value
    return sequence

def plateOrderFromSequence(plateID,sequence):

    #plateID being the corresponding number of the plates for [A,C,G,T]
    plateOrder=[]
    for nuc in sequence:
        if nuc =="A":
            plateOrder.append(plateID[0])
        if nuc =="C":
            plateOrder.append(plateID[1])
        if nuc == "G":
            plateOrder.append(plateID[2])
        if nuc == "T":
            plateOrder.append(plateID[3])
    return plateOrder

def sequenceFromPlateOrder(plateID,plateOrder):
    # plateID being the corresponding number of the plates for [A,C,G,T]
    sequence = ""
    for number in plateOrder:
        index=plateID.index(number)+1
        if index == 1:
            sequence = sequence + "A"
        if index == 2:
            sequence = sequence + "C"
        if index == 3:
            sequence = sequence + "G"
        if index == 4:
            sequence = sequence + "T"


    return sequence

def allMirrorSequences(sequence):

    initialPlateID=[1,2,3,4]
    plateOrder=plateOrderFromSequence(initialPlateID,sequence)
    mirrorSequences=[]
    print("plateOrder = ")
    print(plateOrder)
    seqNb=1

    for Aindex in list(range(1,4+1)):
        Clist=list(range(1,4+1))
        Clist.remove(Aindex)
        for Cindex in Clist:
            Glist=list(range(1,4+1))
            Glist.remove(Aindex)
            Glist.remove(Cindex)
            for Gindex in Glist:
                Tlist = list(range(1, 4 + 1))
                Tlist.remove(Aindex)
                Tlist.remove(Cindex)
                Tlist.remove(Gindex)
                for Tindex in Tlist:
                    print("-------------------------------")
                    print("Mirror Sequence #" + str(seqNb) )
                    print("A = " + str(Aindex) +", C = " + str(Cindex) +", G = " + str(Gindex) +", T = " + str(Tindex))
                    plateID=[Aindex,Cindex,Gindex,Tindex]
                    print(plateID)
                    print(sequenceFromPlateOrder(plateID,plateOrder))
                    mirrorSequences.append(sequenceFromPlateOrder(plateID,plateOrder))
                    seqNb=seqNb+1
    return mirrorSequences


def writeMirrorSequencesExcel(plateOrder,mirrorSequences):
    Mirror_excel = openpyxl.load_workbook(outputPath)
    Mirror_sheet = Mirror_excel["Mirror"]

    # write plateOrder in Excel output
    plateOrderIndexes = []
    for row in range(1, 100):
        for col in range(1, 100):
            value = Mirror_sheet.cell(row, col).value
            if value == "Plate order":
                plateOrderIndexes = [row, col]

    # erase previous platOrder
    for col in range(1, 60):
        Mirror_sheet.cell(plateOrderIndexes[0], plateOrderIndexes[1] + col).value = ""

    # write plateOrder in Excel output
    col = 1
    for plateNbr in plateOrder:
        Mirror_sheet.cell(plateOrderIndexes[0], plateOrderIndexes[1] + col).value = plateNbr
        col = col + 1

    #Write mirror sequences
    mirrorSequencesIndexes = []
    for row in range(1, 100):
        for col in range(1, 100):
            value = Mirror_sheet.cell(row, col).value
            if value == "Mirror sequences":
                mirrorSequencesIndexes = [row, col]
    row = 1
    for seq in mirrorSequences:
        Mirror_sheet.cell(mirrorSequencesIndexes[0] + row, mirrorSequencesIndexes[1] + 1).value = seq
        row = row + 1

    Mirror_excel.save(outputPath)

plateID=[1,2,3,4]
sequence=getMotherSequences(path)
print(sequence)
plateOrder=plateOrderFromSequence(plateID,sequence)
mirrorSequences = allMirrorSequences(sequence)
writeMirrorSequencesExcel(plateOrder,mirrorSequences)

